<<<<<<< HEAD
import logging
from datetime import datetime
from io import BytesIO

import openpyxl
import pandas as pd
import pytz
from common.enums.vacation import VacationStatus
from common.service.bima_service import BimaService
from common.service.template_notification_service import BimaTemplateNotificationService
from company.models import BimaCompany
from core.notification_template.models import BimaCoreNotificationTemplate
from django.apps import apps
from django.db import models
from django.db import transaction, IntegrityError
from django.db.models import Value, CharField
from django.db.models.functions import Concat
from django.utils import timezone
from django.utils.translation import gettext_lazy as _
from hr.vacation.models import BimaHrVacation
from openpyxl.styles import Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


def calculate_vacation_balances(employee):
    bima_company = BimaCompany.objects.first()  # TODO: Handle case where BimaCompany does not exist
    vacation_coefficient = bima_company.vacation_coefficient
    hiring_date = employee.hiring_date if employee.hiring_date else datetime.now()
    current_year = datetime.now().year
    end_of_year = datetime(current_year, 12, 31)
    current_balance = 0
    if hiring_date.year == current_year:
        worked_months = datetime.now().month - hiring_date.month + 1
        current_balance = worked_months * vacation_coefficient
        employee.balance_vacation = current_balance

    elif hiring_date.year < current_year:
        current_balance = datetime.now().month * vacation_coefficient
        employee.balance_vacation = current_balance

    vacation_from_total_working_days_upcoming = sum(
        BimaService.working_days_count(
            vacation.date_start, vacation.date_end,
            bima_company.start_working_day, bima_company.end_working_day
        )
        for vacation in BimaHrVacation.objects.filter(
            employee=employee,
            date_start__lte=datetime.now(),
            status__in=[VacationStatus.APPROVED.name]
        )
    )
    new_balance_vacation = employee.balance_vacation - vacation_from_total_working_days_upcoming
    employee.balance_vacation = round(new_balance_vacation, 2)

    last_vacation_date = BimaHrVacation.objects.filter(
        employee=employee,
        date_end__lte=end_of_year,
        status__in=[VacationStatus.APPROVED.name]
    ).aggregate(models.Max('date_end'))['date_end__max']

    if last_vacation_date:
        months_until_last_vacation = last_vacation_date.month - datetime.now().month
        expected_new_balance = months_until_last_vacation * vacation_coefficient

        total_working_days_upcoming = sum(
            BimaService.working_days_count(
                vacation.date_start, vacation.date_end,
                bima_company.start_working_day, bima_company.end_working_day
            )
            for vacation in BimaHrVacation.objects.filter(
                employee=employee,
                date_start__lte=last_vacation_date,
                status__in=[VacationStatus.APPROVED.name]
            )
        )
    else:
        expected_new_balance = 0
        total_working_days_upcoming = 0

    new_virtual_balance = current_balance + expected_new_balance - total_working_days_upcoming
    employee.virtual_balance_vacation = round(new_virtual_balance, 2)

    employee.save()


def is_vacation_request_valid(vacation):
    bima_company = BimaCompany.objects.first()
    requested_working_days = BimaService.working_days_count(vacation.date_start, vacation.date_end,
                                                            bima_company.start_working_day,
                                                            bima_company.end_working_day)
    return requested_working_days <= vacation.employee.balance_vacation, requested_working_days


def update_vacation_status(vacation, status_update, reason_refused=None, save=True):
    if status_update == VacationStatus.APPROVED.name:
        vacation.status = VacationStatus.APPROVED.name
    elif status_update == VacationStatus.REFUSED.name and reason_refused:
        vacation.status = VacationStatus.REFUSED.name
        vacation.reason_refused = reason_refused
    else:
        raise ValueError('Invalid data')

    vacation.status_change_date = timezone.now()
    if save:
        vacation.save()
    calculate_vacation_balances(vacation.employee)
    return vacation


def update_expired_vacations():
    try:
        with transaction.atomic():
            today = timezone.localdate()
            pending_vacations = BimaHrVacation.objects.filter(status=VacationStatus.PENDING.name)
            updated_count = 0
            for vacation in pending_vacations:
                if vacation.date_end < today:
                    vacation.status = VacationStatus.EXPIRED.name
                    vacation.save()
                    updated_count += 1

            logger.info(f'Successfully updated {updated_count} vacation(s) to expired status.')

    except IntegrityError as e:
        logger.error(f'Database error while updating expired vacations: {e}')

    except Exception as e:
        logger.error(f'Unexpected error while updating expired vacations: {e}')


class BimaHrVacationExportService:

    def __init__(self, queryset):
        self.queryset = queryset
        print(pytz)

    @staticmethod
    def remove_timezone(series):
        return series.dt.tz_localize(None)

    def to_dataframe(self):
        qs_values = self.queryset.annotate(
            employee_full_name=Concat(
                'employee__first_name', Value(' '), 'employee__last_name',
                output_field=CharField()
            ),
            manager_full_name=Concat(
                'manager__first_name', Value(' '), 'manager__last_name',
                output_field=CharField()
            )
        ).values(
            'employee_full_name', 'employee__balance_vacation', 'employee__virtual_balance_vacation',
            'manager_full_name', 'date_start', 'date_end',
            'vacation_type', 'status',
            'request_date', 'status_change_date', 'reason_refused'
        )

        df = pd.DataFrame.from_records(qs_values)

        # Specify column order
        columns_order = [
            'employee_full_name', 'employee__balance_vacation', 'employee__virtual_balance_vacation',
            'manager_full_name', 'date_start', 'date_end',
            'vacation_type', 'status',
            'request_date', 'status_change_date', 'reason_refused'
        ]
        df = df[columns_order]

        datetime_cols = [col for col in df.select_dtypes(include=['datetime64[ns, UTC]']).columns]
        df[datetime_cols] = df[datetime_cols].apply(self.remove_timezone)

        return df

    def translate_enum_values(self, df):
        vacation_type_mapping = {
            'ANNUAL': _('Annual'),
            'SICK': _('Sick'),
            'UNPAID': _('Unpaid'),
            'OTHER': _('Other')
        }
        status_mapping = {
            'PENDING': _('Pending'),
            'APPROVED': _('Approved'),
            'REFUSED': _('Refused')
        }
        df['vacation_type'] = df['vacation_type'].map(vacation_type_mapping)
        df['status'] = df['status'].map(status_mapping)
        return df

    def export_to_csv(self):
        df = self.to_dataframe()
        df = self.translate_enum_values(df)
        df.columns = [
            _('Employee'), _('Vacation Balance'), _('Virtual Vacation Balance'),
            _('Manager'), _('Start Date'), _('End Date'),
            _('Vacation Type'), _('Status'),
            _('Request Date'), _('Date Approve/Refuse'), _('Reason Refuse')
        ]
        buffer = BytesIO()
        df.to_csv(buffer, index=False)
        return buffer.getvalue()

    def export_to_excel(self):
        df = self.to_dataframe()
        df = self.translate_enum_values(df)
        df.columns = [
            _('Employee'), _('Vacation Balance'), _('Virtual Vacation Balance'),
            _('Manager'), _('Start Date'), _('End Date'),
            _('Vacation Type'), _('Status'),
            _('Request Date'), _('Date Approve/Refuse'), _('Reason Refuse')
        ]
        buffer = BytesIO()
        df.to_excel(buffer, index=False, engine='openpyxl')
        buffer.seek(0)
        return buffer.getvalue()

    def export_employee_vacation_to_excel(self, employee=None):
        buffer = BytesIO()

        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            if employee:
                data = {
                    _('Employee'): [employee.full_name],
                    _('Vacation Balance'): [employee.balance_vacation],
                    _('Vacation Virtual Balance'): [employee.virtual_balance_vacation],
                }
                df = pd.DataFrame(data)
                df.to_excel(writer, index=False, sheet_name='Employee Info')

                vacations = employee.bimahrvacation_set.all().values('date_start', 'date_end', 'status', 'request_date',
                                                                     'status_change_date')
                df_vacations = pd.DataFrame.from_records(vacations)
                df_vacations.columns = [_('start_date'), _('end_date'), _('status'), _('request_date'),
                                        _('date_approve_refuse')]
                df_vacations.to_excel(writer, index=False, sheet_name='Vacations')

            else:
                BimaHrEmployee = apps.get_model('hr', 'BimaHrEmployee')
                employees = BimaHrEmployee.objects.all()
                data = {
                    _('Employee'): [e.full_name for e in employees],
                    _('Vacation Balance'): [e.balance_vacation for e in employees],
                    _('Vacation Virtual Balance'): [e.virtual_balance_vacation for e in employees],
                }
                df = pd.DataFrame(data)
                df.to_excel(writer, index=False, sheet_name='Employees Info')

            workbook = writer.book

            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.border = thin_border

        buffer.seek(0)
        return buffer.getvalue()


class ExcelExporter:
    def __init__(self):
        self.wb = openpyxl.Workbook()
        self.border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        self.header_font = Font(bold=True)
        self.header_fill = PatternFill(start_color="FFC0C0C0",
                                       end_color="FFC0C0C0", fill_type="solid")

    def apply_border(self, sheet, cell_range):
        rows = list(sheet[cell_range])
        for row in rows:
            for cell in row:
                cell.border = self.border

    def apply_header_style(self, sheet, cell_range):
        rows = list(sheet[cell_range])
        for row in rows:
            for cell in row:
                cell.font = self.header_font
                cell.fill = self.header_fill

    def auto_adjust_column_width(self, sheet):
        for col in sheet.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            sheet.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    def save(self, buffer):
        self.wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()


class EmployeeExporter(ExcelExporter):
    @staticmethod
    def remove_timezone(dt):
        return dt.replace(tzinfo=None) if dt else None

    def __init__(self, employee):
        super().__init__()
        self.employee = employee

    def export(self):
        sheet = self.wb.active
        sheet.title = 'Employee Info'

        if self.employee:
            self._export_single_employee(sheet)
        else:
            self._export_all_employees(sheet)

    def _format_employee_data(self, employee):
        return [
            ['Unique ID', employee.unique_id],
            ['First Name', employee.first_name],
            ['Last Name', employee.last_name],
            ['Date of Birth',
             employee.date_of_birth.strftime('%Y-%m-%d') if employee.date_of_birth else None],
            ['Place of Birth', employee.place_of_birth],
            ['Country', str(employee.country) if employee.country else None],
            ['Nationality', employee.nationality],
            ['Identity Card Number', employee.identity_card_number],
            ['Phone Number', employee.phone_number],
            ['Second Phone Number', employee.second_phone_number],
            ['Email', employee.email],
            ['Gender', employee.gender],
            ['Marital Status', employee.marital_status],
            ['Number of Children', employee.num_children],
            ['Education Level', employee.education_level],
            ['Latest Degree', employee.latest_degree],
            ['Latest Degree Date',
             employee.latest_degree_date.strftime('%Y-%m-%d') if employee.latest_degree_date else None],
            ['Institute', employee.institute],
            ['Employment Type', employee.employment_type],
            ['Work Mode', employee.work_mode],
            ['Job Type', employee.job_type],
            ['Employment Status', employee.employment_status],
            ['Hiring Date', employee.hiring_date.strftime('%Y-%m-%d') if employee.hiring_date else None],
            ['Probation End Date',
             employee.probation_end_date.strftime('%Y-%m-%d') if employee.probation_end_date else None],
            ['Last Performance Review', employee.last_performance_review.strftime(
                '%Y-%m-%d') if employee.last_performance_review else None],
            ['Salary', employee.salary],
            ['Position', str(employee.position) if employee.position else None],
            ['Balance Vacation', employee.balance_vacation],
            ['Virtual Balance Vacation', employee.virtual_balance_vacation],
        ]

    def _export_single_employee(self, sheet):
        # Employee details title
        sheet.append(['Employee Details'])
        self.apply_header_style(sheet, 'A1:A1')

        # Employee details
        employee_details = self._format_employee_data(self.employee)
        for row in employee_details:
            sheet.append(row)
        self.apply_border(sheet, f'A2:B{len(employee_details) + 1}')

        # Vacation balance title
        sheet.append(['', ''])  # Empty row
        sheet.append(['Employee Vacation Balance'])
        self.apply_header_style(sheet, f'A{len(employee_details) + 3}:A{len(employee_details) + 3}')

        # Vacation balance
        vacation_balance = [
            ['Vacation Balance', self.employee.balance_vacation],
            ['Vacation Virtual Balance', self.employee.virtual_balance_vacation],
        ]
        for row in vacation_balance:
            sheet.append(row)
        self.apply_border(sheet, f'A{len(employee_details) + 4}:B{len(employee_details) + 5}')

        # Employee vacations title
        sheet.append(['', ''])  # Empty row
        sheet.append(['Employee Vacation Details'])
        self.apply_header_style(sheet, f'A{len(employee_details) + 7}:A{len(employee_details) + 7}')

        # Employee vacations
        vacations = self.employee.bimahrvacation_set.all().values(
            'date_start', 'date_end', 'status', 'request_date', 'status_change_date'
        )
        sheet.append(['Start Date', 'End Date', 'Status', 'Request Date', 'Status Change Date'])
        for vacation in vacations:
            # Remove timezone info from datetime fields
            vacation['request_date'] = self.remove_timezone(vacation['request_date'])
            vacation['status_change_date'] = self.remove_timezone(vacation['status_change_date'])
            sheet.append([vacation[field] for field in vacations[0].keys()])

        self.apply_border(sheet, f'A{len(employee_details) + 8}:E{len(employee_details) + 8 + len(vacations)}')

        self.auto_adjust_column_width(sheet)

    def _employee_to_list(self, employee):
        # Convert an employee to a list of data
        formatted_data = self._format_employee_data(employee)
        return [data[1] for data in formatted_data]  # Return only the values

    def _export_all_employees(self, sheet):
        from django.apps import apps

        BimaHrEmployee = apps.get_model('hr', 'BimaHrEmployee')
        employees = BimaHrEmployee.objects.all()

        # Header
        header = [data[0] for data in
                  self._format_employee_data(self.employee or BimaHrEmployee())]  # Return only the keys
        sheet.append(header)
        self.apply_header_style(sheet, f'A1:AB1')

        for employee in employees:
            employee_data = self._employee_to_list(employee)
            sheet.append(employee_data)

        self.apply_border(sheet, f'A2:AB{len(employees) + 1}')
        self.auto_adjust_column_width(sheet)


class BimaVacationNotificationService:

    @staticmethod
    def send_notification_request_vacation(vacation):
        notification_template, data_to_send = BimaVacationNotificationService.get_and_format_template(
            'NOTIFICATION_VACATION_REQUEST', vacation)
        data = BimaVacationNotificationService.prepare_data(notification_template, vacation, data_to_send)
        BimaTemplateNotificationService.send_email_and_save_notification.delay(data)

    @staticmethod
    def send_approval_refusal_notification(vacation):
        notification_code = 'NOTIFICATION_VACATION_APPROVAL' if vacation.status == VacationStatus.APPROVED.name \
            else 'NOTIFICATION_VACATION_REFUSAL'
        notification_template, data_to_send = BimaVacationNotificationService.get_and_format_template(
            notification_code, vacation
        )
        data = BimaVacationNotificationService.prepare_data(notification_template, vacation, data_to_send,
                                                            for_manager=False)
        BimaTemplateNotificationService.send_email_and_save_notification.delay(data)

    @staticmethod
    def prepare_data(notification_template, vacation, data_to_send, for_manager=True):
        return {
            'subject': data_to_send['subject'],
            'message': data_to_send['message'],
            'receivers': [vacation.manager.email if for_manager else vacation.employee.email],
            'attachments': [] if data_to_send['file_url'] is None else [data_to_send['file_url']],
            'notification_type_id': notification_template.notification_type.id,
            'sender': None,
            'app_name': 'hr',
            'model_name': 'bimahrvacation',
            'parent_id': vacation.id
        }

    @staticmethod
    def get_and_format_template(notification_code, vacation):
        template = BimaCoreNotificationTemplate.objects.filter(
            notification_type__code=notification_code
        ).first()
        if not template:
            raise ValueError("Notification template not found")

        data = {
            'manager_full_name': vacation.manager.full_name,
            'employee_full_name': vacation.employee.full_name,
            'vacation_date_start': vacation.date_start,
            'vacation_date_end': vacation.date_end,
            'vacation_reason': vacation.reason,
            'reason_refused': vacation.reason_refused or ''
        }

        formatted_message = BimaTemplateNotificationService.replace_variables_in_template(
            template.message, data
        )
        formatted_subject = BimaTemplateNotificationService.replace_variables_in_template(
            template.subject, data
        )
        return template, {'subject': formatted_subject, 'message': formatted_message, 'file_url': None}
=======
import logging
from datetime import datetime
from io import BytesIO

import openpyxl
import pandas as pd
import pytz
from common.enums.vacation import VacationStatus
from common.service.bima_service import BimaService
from common.service.template_notification_service import BimaTemplateNotificationService
from company.models import BimaCompany
from core.notification_template.models import BimaCoreNotificationTemplate
from django.apps import apps
from django.db import models
from django.db import transaction, IntegrityError
from django.db.models import Value, CharField
from django.db.models.functions import Concat
from django.utils import timezone
from django.utils.translation import gettext_lazy as _
from hr.vacation.models import BimaHrVacation
from openpyxl.styles import Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


def calculate_vacation_balances(employee):
    bima_company = BimaCompany.objects.first()  # TODO: Handle case where BimaCompany does not exist
    vacation_coefficient = bima_company.vacation_coefficient
    hiring_date = employee.hiring_date if employee.hiring_date else datetime.now()
    current_year = datetime.now().year
    end_of_year = datetime(current_year, 12, 31)
    current_balance = 0
    if hiring_date.year == current_year:
        worked_months = datetime.now().month - hiring_date.month + 1
        current_balance = worked_months * vacation_coefficient
        employee.balance_vacation = current_balance

    elif hiring_date.year < current_year:
        current_balance = datetime.now().month * vacation_coefficient
        employee.balance_vacation = current_balance

    vacation_from_total_working_days_upcoming = sum(
        BimaService.working_days_count(
            vacation.date_start, vacation.date_end,
            bima_company.start_working_day, bima_company.end_working_day
        )
        for vacation in BimaHrVacation.objects.filter(
            employee=employee,
            date_start__lte=datetime.now(),
            status__in=[VacationStatus.APPROVED.name]
        )
    )
    new_balance_vacation = employee.balance_vacation - vacation_from_total_working_days_upcoming
    employee.balance_vacation = round(new_balance_vacation, 2)

    last_vacation_date = BimaHrVacation.objects.filter(
        employee=employee,
        date_end__lte=end_of_year,
        status__in=[VacationStatus.APPROVED.name]
    ).aggregate(models.Max('date_end'))['date_end__max']

    if last_vacation_date:
        months_until_last_vacation = last_vacation_date.month - datetime.now().month
        expected_new_balance = months_until_last_vacation * vacation_coefficient

        total_working_days_upcoming = sum(
            BimaService.working_days_count(
                vacation.date_start, vacation.date_end,
                bima_company.start_working_day, bima_company.end_working_day
            )
            for vacation in BimaHrVacation.objects.filter(
                employee=employee,
                date_start__lte=last_vacation_date,
                status__in=[VacationStatus.APPROVED.name]
            )
        )
    else:
        expected_new_balance = 0
        total_working_days_upcoming = 0

    new_virtual_balance = current_balance + expected_new_balance - total_working_days_upcoming
    employee.virtual_balance_vacation = round(new_virtual_balance, 2)

    employee.save()


def is_vacation_request_valid(vacation):
    bima_company = BimaCompany.objects.first()
    requested_working_days = BimaService.working_days_count(vacation.date_start, vacation.date_end,
                                                            bima_company.start_working_day,
                                                            bima_company.end_working_day)
    return requested_working_days <= vacation.employee.balance_vacation, requested_working_days


def update_vacation_status(vacation, status_update, reason_refused=None, save=True):
    if status_update == VacationStatus.APPROVED.name:
        vacation.status = VacationStatus.APPROVED.name
    elif status_update == VacationStatus.REFUSED.name and reason_refused:
        vacation.status = VacationStatus.REFUSED.name
        vacation.reason_refused = reason_refused
    else:
        raise ValueError('Invalid data')

    vacation.status_change_date = timezone.now()
    if save:
        vacation.save()
    calculate_vacation_balances(vacation.employee)
    return vacation


def update_expired_vacations():
    try:
        with transaction.atomic():
            today = timezone.localdate()
            pending_vacations = BimaHrVacation.objects.filter(status=VacationStatus.PENDING.name)
            updated_count = 0
            for vacation in pending_vacations:
                if vacation.date_end < today:
                    vacation.status = VacationStatus.EXPIRED.name
                    vacation.save()
                    updated_count += 1

            logger.info(f'Successfully updated {updated_count} vacation(s) to expired status.')

    except IntegrityError as e:
        logger.error(f'Database error while updating expired vacations: {e}')

    except Exception as e:
        logger.error(f'Unexpected error while updating expired vacations: {e}')


class BimaHrVacationExportService:

    def __init__(self, queryset):
        self.queryset = queryset
        print(pytz)

    @staticmethod
    def remove_timezone(series):
        return series.dt.tz_localize(None)

    def to_dataframe(self):
        qs_values = self.queryset.annotate(
            employee_full_name=Concat(
                'employee__first_name', Value(' '), 'employee__last_name',
                output_field=CharField()
            ),
            manager_full_name=Concat(
                'manager__first_name', Value(' '), 'manager__last_name',
                output_field=CharField()
            )
        ).values(
            'employee_full_name', 'employee__balance_vacation', 'employee__virtual_balance_vacation',
            'manager_full_name', 'date_start', 'date_end',
            'vacation_type', 'status',
            'request_date', 'status_change_date', 'reason_refused'
        )

        df = pd.DataFrame.from_records(qs_values)

        # Specify column order
        columns_order = [
            'employee_full_name', 'employee__balance_vacation', 'employee__virtual_balance_vacation',
            'manager_full_name', 'date_start', 'date_end',
            'vacation_type', 'status',
            'request_date', 'status_change_date', 'reason_refused'
        ]
        df = df[columns_order]

        datetime_cols = [col for col in df.select_dtypes(include=['datetime64[ns, UTC]']).columns]
        df[datetime_cols] = df[datetime_cols].apply(self.remove_timezone)

        return df

    def translate_enum_values(self, df):
        vacation_type_mapping = {
            'ANNUAL': _('Annual'),
            'SICK': _('Sick'),
            'UNPAID': _('Unpaid'),
            'OTHER': _('Other')
        }
        status_mapping = {
            'PENDING': _('Pending'),
            'APPROVED': _('Approved'),
            'REFUSED': _('Refused')
        }
        df['vacation_type'] = df['vacation_type'].map(vacation_type_mapping)
        df['status'] = df['status'].map(status_mapping)
        return df

    def export_to_csv(self):
        df = self.to_dataframe()
        df = self.translate_enum_values(df)
        df.columns = [
            _('Employee'), _('Vacation Balance'), _('Virtual Vacation Balance'),
            _('Manager'), _('Start Date'), _('End Date'),
            _('Vacation Type'), _('Status'),
            _('Request Date'), _('Date Approve/Refuse'), _('Reason Refuse')
        ]
        buffer = BytesIO()
        df.to_csv(buffer, index=False)
        return buffer.getvalue()

    def export_to_excel(self):
        df = self.to_dataframe()
        df = self.translate_enum_values(df)
        df.columns = [
            _('Employee'), _('Vacation Balance'), _('Virtual Vacation Balance'),
            _('Manager'), _('Start Date'), _('End Date'),
            _('Vacation Type'), _('Status'),
            _('Request Date'), _('Date Approve/Refuse'), _('Reason Refuse')
        ]
        buffer = BytesIO()
        df.to_excel(buffer, index=False, engine='openpyxl')
        buffer.seek(0)
        return buffer.getvalue()

    def export_employee_vacation_to_excel(self, employee=None):
        buffer = BytesIO()

        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            if employee:
                data = {
                    _('Employee'): [employee.full_name],
                    _('Vacation Balance'): [employee.balance_vacation],
                    _('Vacation Virtual Balance'): [employee.virtual_balance_vacation],
                }
                df = pd.DataFrame(data)
                df.to_excel(writer, index=False, sheet_name='Employee Info')

                vacations = employee.bimahrvacation_set.all().values('date_start', 'date_end', 'status', 'request_date',
                                                                     'status_change_date')
                df_vacations = pd.DataFrame.from_records(vacations)
                df_vacations.columns = [_('start_date'), _('end_date'), _('status'), _('request_date'),
                                        _('date_approve_refuse')]
                df_vacations.to_excel(writer, index=False, sheet_name='Vacations')

            else:
                BimaHrEmployee = apps.get_model('hr', 'BimaHrEmployee')
                employees = BimaHrEmployee.objects.all()
                data = {
                    _('Employee'): [e.full_name for e in employees],
                    _('Vacation Balance'): [e.balance_vacation for e in employees],
                    _('Vacation Virtual Balance'): [e.virtual_balance_vacation for e in employees],
                }
                df = pd.DataFrame(data)
                df.to_excel(writer, index=False, sheet_name='Employees Info')

            workbook = writer.book

            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.border = thin_border

        buffer.seek(0)
        return buffer.getvalue()


class ExcelExporter:
    def __init__(self):
        self.wb = openpyxl.Workbook()
        self.border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        self.header_font = Font(bold=True)
        self.header_fill = PatternFill(start_color="FFC0C0C0",
                                       end_color="FFC0C0C0", fill_type="solid")

    def apply_border(self, sheet, cell_range):
        rows = list(sheet[cell_range])
        for row in rows:
            for cell in row:
                cell.border = self.border

    def apply_header_style(self, sheet, cell_range):
        rows = list(sheet[cell_range])
        for row in rows:
            for cell in row:
                cell.font = self.header_font
                cell.fill = self.header_fill

    def auto_adjust_column_width(self, sheet):
        for col in sheet.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            sheet.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    def save(self, buffer):
        self.wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()


class EmployeeExporter(ExcelExporter):
    @staticmethod
    def remove_timezone(dt):
        return dt.replace(tzinfo=None) if dt else None

    def __init__(self, employee):
        super().__init__()
        self.employee = employee

    def export(self):
        sheet = self.wb.active
        sheet.title = 'Employee Info'

        if self.employee:
            self._export_single_employee(sheet)
        else:
            self._export_all_employees(sheet)

    def _format_employee_data(self, employee):
        return [
            ['Unique ID', employee.unique_id],
            ['First Name', employee.first_name],
            ['Last Name', employee.last_name],
            ['Date of Birth',
             employee.date_of_birth.strftime('%Y-%m-%d') if employee.date_of_birth else None],
            ['Place of Birth', employee.place_of_birth],
            ['Country', str(employee.country) if employee.country else None],
            ['Nationality', employee.nationality],
            ['Identity Card Number', employee.identity_card_number],
            ['Phone Number', employee.phone_number],
            ['Second Phone Number', employee.second_phone_number],
            ['Email', employee.email],
            ['Gender', employee.gender],
            ['Marital Status', employee.marital_status],
            ['Number of Children', employee.num_children],
            ['Education Level', employee.education_level],
            ['Latest Degree', employee.latest_degree],
            ['Latest Degree Date',
             employee.latest_degree_date.strftime('%Y-%m-%d') if employee.latest_degree_date else None],
            ['Institute', employee.institute],
            ['Employment Type', employee.employment_type],
            ['Work Mode', employee.work_mode],
            ['Job Type', employee.job_type],
            ['Employment Status', employee.employment_status],
            ['Hiring Date', employee.hiring_date.strftime('%Y-%m-%d') if employee.hiring_date else None],
            ['Probation End Date',
             employee.probation_end_date.strftime('%Y-%m-%d') if employee.probation_end_date else None],
            ['Last Performance Review', employee.last_performance_review.strftime(
                '%Y-%m-%d') if employee.last_performance_review else None],
            ['Salary', employee.salary],
            ['Position', str(employee.position) if employee.position else None],
            ['Balance Vacation', employee.balance_vacation],
            ['Virtual Balance Vacation', employee.virtual_balance_vacation],
        ]

    def _export_single_employee(self, sheet):
        # Employee details title
        sheet.append(['Employee Details'])
        self.apply_header_style(sheet, 'A1:A1')

        # Employee details
        employee_details = self._format_employee_data(self.employee)
        for row in employee_details:
            sheet.append(row)
        self.apply_border(sheet, f'A2:B{len(employee_details) + 1}')

        # Vacation balance title
        sheet.append(['', ''])  # Empty row
        sheet.append(['Employee Vacation Balance'])
        self.apply_header_style(sheet, f'A{len(employee_details) + 3}:A{len(employee_details) + 3}')

        # Vacation balance
        vacation_balance = [
            ['Vacation Balance', self.employee.balance_vacation],
            ['Vacation Virtual Balance', self.employee.virtual_balance_vacation],
        ]
        for row in vacation_balance:
            sheet.append(row)
        self.apply_border(sheet, f'A{len(employee_details) + 4}:B{len(employee_details) + 5}')

        # Employee vacations title
        sheet.append(['', ''])  # Empty row
        sheet.append(['Employee Vacation Details'])
        self.apply_header_style(sheet, f'A{len(employee_details) + 7}:A{len(employee_details) + 7}')

        # Employee vacations
        vacations = self.employee.bimahrvacation_set.all().values(
            'date_start', 'date_end', 'status', 'request_date', 'status_change_date'
        )
        sheet.append(['Start Date', 'End Date', 'Status', 'Request Date', 'Status Change Date'])
        for vacation in vacations:
            # Remove timezone info from datetime fields
            vacation['request_date'] = self.remove_timezone(vacation['request_date'])
            vacation['status_change_date'] = self.remove_timezone(vacation['status_change_date'])
            sheet.append([vacation[field] for field in vacations[0].keys()])

        self.apply_border(sheet, f'A{len(employee_details) + 8}:E{len(employee_details) + 8 + len(vacations)}')

        self.auto_adjust_column_width(sheet)

    def _employee_to_list(self, employee):
        # Convert an employee to a list of data
        formatted_data = self._format_employee_data(employee)
        return [data[1] for data in formatted_data]  # Return only the values

    def _export_all_employees(self, sheet):
        from django.apps import apps

        BimaHrEmployee = apps.get_model('hr', 'BimaHrEmployee')
        employees = BimaHrEmployee.objects.all()

        # Header
        header = [data[0] for data in
                  self._format_employee_data(self.employee or BimaHrEmployee())]  # Return only the keys
        sheet.append(header)
        self.apply_header_style(sheet, f'A1:AB1')

        for employee in employees:
            employee_data = self._employee_to_list(employee)
            sheet.append(employee_data)

        self.apply_border(sheet, f'A2:AB{len(employees) + 1}')
        self.auto_adjust_column_width(sheet)


class BimaVacationNotificationService:

    @staticmethod
    def send_notification_request_vacation(vacation):
        notification_template, data_to_send = BimaVacationNotificationService.get_and_format_template(
            'NOTIFICATION_VACATION_REQUEST', vacation)
        data = BimaVacationNotificationService.prepare_data(notification_template, vacation, data_to_send)
        BimaTemplateNotificationService.send_email_and_save_notification.delay(data)

    @staticmethod
    def send_approval_refusal_notification(vacation):
        notification_code = 'NOTIFICATION_VACATION_APPROVAL' if vacation.status == VacationStatus.APPROVED.name \
            else 'NOTIFICATION_VACATION_REFUSAL'
        notification_template, data_to_send = BimaVacationNotificationService.get_and_format_template(
            notification_code, vacation
        )
        data = BimaVacationNotificationService.prepare_data(notification_template, vacation, data_to_send,
                                                            for_manager=False)
        BimaTemplateNotificationService.send_email_and_save_notification.delay(data)

    @staticmethod
    def prepare_data(notification_template, vacation, data_to_send, for_manager=True):
        return {
            'subject': data_to_send['subject'],
            'message': data_to_send['message'],
            'receivers': [vacation.manager.email if for_manager else vacation.employee.email],
            'attachments': [] if data_to_send['file_url'] is None else [data_to_send['file_url']],
            'notification_type_id': notification_template.notification_type.id,
            'sender': None,
            'app_name': 'hr',
            'model_name': 'bimahrvacation',
            'parent_id': vacation.id
        }

    @staticmethod
    def get_and_format_template(notification_code, vacation):
        template = BimaCoreNotificationTemplate.objects.filter(
            notification_type__code=notification_code
        ).first()
        if not template:
            raise ValueError("Notification template not found")

        data = {
            'manager_full_name': vacation.manager.full_name,
            'employee_full_name': vacation.employee.full_name,
            'vacation_date_start': vacation.date_start,
            'vacation_date_end': vacation.date_end,
            'vacation_reason': vacation.reason,
            'reason_refused': vacation.reason_refused or ''
        }

        formatted_message = BimaTemplateNotificationService.replace_variables_in_template(
            template.message, data
        )
        formatted_subject = BimaTemplateNotificationService.replace_variables_in_template(
            template.subject, data
        )
        return template, {'subject': formatted_subject, 'message': formatted_message, 'file_url': None}
>>>>>>> origin/ma-branch
