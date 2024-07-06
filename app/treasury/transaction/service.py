import logging
from collections import defaultdict
from datetime import datetime, date, timedelta
from decimal import Decimal

import pandas as pd
from common.enums.transaction_enum import TransactionDirection, TransactionNature
from common.enums.transaction_enum import TransactionTypeIncomeOutcome
from common.helpers.bima360_helper import Bima360Helper
from django.apps import apps
from django.db import models
from django.db.models import Sum, Case, When, F, Avg, Q
from django.utils import timezone
from django.utils.crypto import get_random_string
from django.utils.translation import gettext_lazy as _
from erp.partner.models import BimaErpPartner
from openpyxl.styles import Border, Side, Font
from openpyxl.utils import get_column_letter
from rest_framework.exceptions import ValidationError
from treasury.bank_account.models import BimaTreasuryBankAccount
from treasury.cash.models import BimaTreasuryCash
from treasury.payment_method.models import BimaTreasuryPaymentMethod
from treasury.transaction_type.models import BimaTreasuryTransactionType

logger = logging.getLogger(__name__)


class BimaTreasuryTransactionService:
    def __init__(self, queryset):
        self.queryset = queryset

    def calculate_sums(self):
        aggregations = self.queryset.aggregate(
            total_income=Sum(
                Case(
                    When(direction=TransactionDirection.INCOME.name, then=F("amount")),
                    output_field=models.DecimalField(),
                )
            ),
            total_outcome=Sum(
                Case(
                    When(direction=TransactionDirection.OUTCOME.name, then=F("amount")),
                    output_field=models.DecimalField(),
                )
            ),
        )

        total_income = aggregations.get("total_income") or 0
        total_outcome = aggregations.get("total_outcome") or 0
        difference = total_income - total_outcome

        return {
            "total_income": total_income,
            "total_outcome": total_outcome,
            "difference": difference,
        }

    def format_data_for_export(self):
        data = []
        for transaction in self.queryset:
            try:
                row = {
                    "Number": transaction.number,
                    "Nature": transaction.nature,
                    "Direction": transaction.direction,
                    "Transaction Type": transaction.transaction_type.name,
                    "Transaction Payment Method": transaction.payment_method.name if transaction.payment_method else "",
                    "Note": transaction.note,
                    "Date": transaction.date,
                    "Expected Date": transaction.expected_date,
                    "Amount": transaction.amount,
                    "Transaction Source": transaction.transaction_source.number if transaction.transaction_source else "",
                    "Bank Account Number": transaction.partner_bank_account_number,
                    "Partner": self._get_partner_name(transaction.partner),
                    "Cash": transaction.cash.name if transaction.cash else "",
                    "Bank Account": self._get_bank_account_name(
                        transaction.bank_account
                    ),
                }
                data.append(row)
            except Exception as e:
                logger.error(
                    f"Error processing transaction {transaction.public_id}. Error: {e}"
                )
        return data

    def _get_partner_name(self, partner):
        if partner:
            if partner.partner_type == "INDIVIDUAL":
                return f"{partner.first_name} {partner.last_name}"
            else:
                return partner.company_name
        return ""

    def _get_bank_account_name(self, bank_account):
        if bank_account:
            return f"{bank_account.name} ({bank_account.account_number})"
        return ""

    def export_to_csv(self):
        data = self.format_data_for_export()
        return pd.DataFrame(data)

    def export_to_excel(self):
        data = self.format_data_for_export()
        return pd.DataFrame(data)

    @staticmethod
    def create_auto_transaction(transaction):
        from .models import BimaTreasuryTransaction

        complementary_transaction_mappings = {
            ("FROM_CASH_TO_ACCOUNT_OUTCOME", "OUTCOME"): (
                "FROM_CASH_TO_ACCOUNT_INCOME",
                "INCOME",
            ),
            ("FROM_ACCOUNT_TO_CASH_OUTCOME", "OUTCOME"): (
                "FROM_ACCOUNT_TO_CASH_INCOME",
                "INCOME",
            ),
        }

        key = (transaction.transaction_type.code, transaction.direction)
        if key not in complementary_transaction_mappings:
            return

        (
            complementary_code,
            complementary_direction,
        ) = complementary_transaction_mappings[key]

        complementary_transaction_type = BimaTreasuryTransactionType.objects.get(
            code=complementary_code, income_outcome=complementary_direction
        )
        payment_method_for_cash = BimaTreasuryPaymentMethod.objects.get(
            code="CASH_PAYMENT_INCOME", income_outcome="INCOME"
        )
        payment_method_for_bank = BimaTreasuryPaymentMethod.objects.get(
            code="TRANSFER_INCOME", income_outcome="INCOME"
        )
        number = BimaTreasuryTransactionService.generate_unique_number()
        if transaction.nature == TransactionNature.CASH.name:
            params = {
                "number": number,
                "nature": TransactionNature.BANK.name,
                "direction": TransactionDirection.INCOME.name,
                "transaction_type": complementary_transaction_type,
                "payment_method": payment_method_for_bank,
                "bank_account": transaction.bank_account,
                "cash": transaction.cash,
                "date": transaction.date,
                "amount": transaction.amount,
                "transaction_source": transaction,
                "reference": transaction.reference,
            }
        elif transaction.nature == TransactionNature.BANK.name:
            params = {
                "number": number,
                "nature": TransactionNature.CASH.name,
                "direction": TransactionDirection.INCOME.name,
                "transaction_type": complementary_transaction_type,
                "payment_method": payment_method_for_cash,
                "cash": transaction.cash,
                "bank_account": transaction.bank_account,
                "date": transaction.date,
                "amount": transaction.amount,
                "transaction_source": transaction,
                "reference": transaction.reference,
            }
        BimaTreasuryTransaction.__class__.skip_child_validation = True
        created_transaction = BimaTreasuryTransaction.objects.create(**params)
        BimaTreasuryTransaction.__class__.skip_child_validation = False
        logger.info(
            f"Auto transaction {created_transaction.pk} created for transaction {transaction.pk}"
        )
        return created_transaction

    @staticmethod
    def group_by_date(history_data):
        grouped = defaultdict(list)

        for record in history_data:
            date_without_time = datetime.fromisoformat(record["history_date"]).date()
            grouped[date_without_time].append(record)

        return grouped

    def get_bank_account_name(self, value):
        obj = BimaTreasuryBankAccount.objects.filter(id=value).first()
        return obj.name if obj else None

    def get_cash_name(self, value):
        obj = BimaTreasuryCash.objects.filter(id=value).first()
        return obj.name if obj else None

    def get_transaction_type_name(self, value):
        obj = BimaTreasuryTransactionType.objects.filter(id=value).first()
        return obj.name if obj else None

    def get_partner_name(self, value):
        obj = BimaErpPartner.objects.filter(id=value).first()
        if obj:
            if obj.partner_type == "INDIVIDUAL":
                return f"{obj.first_name} {obj.last_name}"
            else:
                return obj.company_name
        return None

    FIELD_TO_METHOD_MAP = {
        "bank_account": get_bank_account_name,
        "cash": get_cash_name,
        "transaction_type": get_transaction_type_name,
        "partner": get_partner_name,
    }

    def style_worksheet(self, ws):
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border

        for row in ws.iter_rows():
            for cell in row:
                if cell.column_letter == "B":
                    if cell.value == "INCOME":
                        amount_cell = ws[f"G{cell.row}"]
                        amount_cell.font = Font(color="00FF00")
                    elif cell.value == "OUTCOME":
                        amount_cell = ws[f"G{cell.row}"]
                        amount_cell.font = Font(color="FF0000")

        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[
                get_column_letter(column[0].column)
            ].width = adjusted_width

    def append_totals(self, ws):
        sums = self.calculate_sums()
        ws.append([])
        ws.append(["Totals"])
        ws.append(["Total Income", sums["total_income"]])
        ws.append(["Total Outcome", sums["total_outcome"]])
        ws.append(["Difference", sums["difference"]])

    @staticmethod
    def generate_unique_number():
        first_char = "B"
        second_char = "T"
        year = datetime.now().year
        random_string = get_random_string(length=12, allowed_chars='ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789')
        unique_number = f"{first_char}{second_char}_{year}_{random_string}"
        return unique_number

    @staticmethod
    def verify_if_payment_credit_not_and_verify_amounts_from_request(data_from_request, document_public_id):
        transaction_type_code = data_from_request.get("transaction_type", None)
        amount = data_from_request.get("amount", 0)
        if not transaction_type_code:
            return False
        transaction_type_code = transaction_type_code.code
        if not transaction_type_code in ["CREDIT_NOTE_OUTCOME_BANK", "CREDIT_NOTE_OUTCOME_CASH"]:
            return True
        BimaErpSaleDocument = apps.get_model('erp', 'BimaErpSaleDocument')
        sale_documents = BimaErpSaleDocument.objects.filter(
            public_id__in=document_public_id
        )
        total_sale_document_amounts = sum(sd.total_amount for sd in sale_documents)
        if amount != total_sale_document_amounts:
            raise ValidationError(
                {"Error": _("Amount of the transaction should be equal to the amount of credit note!")})
        return True

    @staticmethod
    def get_top_n_transaction_by_type_direction(number_of_type_transaction, direction, year, month):
        from .models import BimaTreasuryTransaction
        transactions = BimaTreasuryTransaction.objects.filter(
            direction=direction,
            date__year=year,
            date__month=month
        ).values(
            'transaction_type__name'  # Assuming there's a name field in transaction_type model
        ).annotate(
            total_amount=Sum('amount')
        ).order_by(
            '-total_amount'
        )[:number_of_type_transaction]

        result = {}
        for trans in transactions:
            transaction_type_name = trans['transaction_type__name']
            result[transaction_type_name] = trans['total_amount']

        return result

    @staticmethod
    def avg_transaction_by_direction(year):
        from .models import BimaTreasuryTransaction
        year = int(year)
        transactions = BimaTreasuryTransaction.objects.filter(
            date__year=year
        ).values(
            month=F('date__month'),
            transaction_direction=F('direction')
        ).annotate(
            avg_amount=Avg('amount')
        ).order_by('month')

        result = {
            i: {
                TransactionTypeIncomeOutcome.INCOME.name: 0,
                TransactionTypeIncomeOutcome.OUTCOME.name: 0
            }
            for i in range(1, 13)
        }

        for trans in transactions:
            month = trans['month']
            result[month][trans['transaction_direction']] = trans['avg_amount']

        return result

    @staticmethod
    def top_partners_by_month(direction, n, year):
        from .models import BimaTreasuryTransaction
        transactions = BimaTreasuryTransaction.objects.filter(
            direction=direction,
            date__year=year  # Added year filter
        ).values(
            month=F('date__month'),
            contact_first_name=F('partner__first_name'),
            contact_last_name=F('partner__last_name'),
            company_name=F('partner__company_name')
        ).annotate(
            total_amount=Sum('amount')
        ).order_by('month', '-total_amount')

        # Initialize a default result structure for all 12 months
        result = {i: {} for i in range(1, 13)}

        for trans in transactions:
            month = trans['month']
            partner_name = f"{trans['contact_first_name']} {trans['contact_last_name']} {trans['company_name']}"
            result[month][partner_name] = trans['total_amount']

        for month, partners in result.items():
            sorted_partners = dict(sorted(partners.items(), key=lambda item: item[1], reverse=True)[:n])
            result[month] = sorted_partners

        return result

    @staticmethod
    def monthly_totals(year):
        from .models import BimaTreasuryTransaction
        year = int(year)
        transactions = BimaTreasuryTransaction.objects.filter(
            date__year=year
        ).values(
            month=F('date__month'),
            transaction_direction=F('direction')
        ).annotate(
            total_amount=Sum('amount')
        ).order_by('month')

        result = {
            i: {
                'total_income': 0,
                'total_outcome': 0,
                'difference': 0  # Initializing difference as well
            }
            for i in range(1, 13)
        }

        for trans in transactions:
            month = trans['month']
            if trans['transaction_direction'] == TransactionTypeIncomeOutcome.INCOME.name:
                result[month]['total_income'] += trans['total_amount']
            else:
                result[month]['total_outcome'] += trans['total_amount']

        # Calculating difference for each month
        for month in result:
            result[month]['difference'] = result[month]['total_income'] - result[month]['total_outcome']

        return result

    @staticmethod
    def remaining_amount_analysis_by_month(year):
        from .models import BimaTreasuryTransaction
        transactions = BimaTreasuryTransaction.objects.filter(
            date__year=year  # Added year filter
        ).values(
            month=F('date__month'),
        ).annotate(
            total_remaining_amount=Sum('remaining_amount')
        ).order_by('month')

        result = {i: 0 for i in range(1, 13)}

        for trans in transactions:
            month = trans['month']
            result[month] = trans['total_remaining_amount']

        return result

    @staticmethod
    def month_over_month_growth(year):
        from .models import BimaTreasuryTransaction
        transactions = BimaTreasuryTransaction.objects.filter(
            date__year=year
        ).values(
            month=F('date__month'),
            transaction_direction=F('direction')
        ).annotate(
            total_amount=Sum('amount')
        ).order_by('month', 'transaction_direction')

        result = {}
        previous_month_totals = {"INCOME": None, "OUTCOME": None}

        for trans in transactions:
            month = trans['month']
            direction = trans['transaction_direction']
            growth_rate = 0
            if previous_month_totals[direction] is not None:
                if previous_month_totals[direction] == 0:  # To avoid division by zero
                    growth_rate = 0
                else:
                    growth_rate = ((trans['total_amount'] - previous_month_totals[direction])
                                   / previous_month_totals[direction]) * 100

            result_key = f'{month - 1}_to_{month}_growth_rate'
            if result_key not in result:
                result[result_key] = {}
            result[result_key][direction.lower()] = round(growth_rate, 2)  # Round to 2 decimal places

            previous_month_totals[direction] = trans['total_amount']

        for i in range(2,
                       13):  # Starting from February (month 2) as January doesn't have a previous month in the same year
            for direction in ["INCOME", "OUTCOME"]:
                result_key = f'{i - 1}_to_{i}_growth_rate'
                if result_key not in result:
                    result[result_key] = {}
                result[result_key].setdefault(direction.lower(), 0)

        # Sort results by month
        result = dict(sorted(result.items(), key=lambda item: int(item[0].split('_to_')[0])))

        return result

    @staticmethod
    def cash_bank_flow_kpi(duration_years, cash_ids, bank_ids):
        from .models import BimaTreasuryTransaction

        end_date = timezone.now().date()
        start_date = date(end_date.year - duration_years + 1, 1, 1)

        q_objects = Q()

        if cash_ids:
            q_objects |= Q(cash__public_id__in=cash_ids)

        if bank_ids:
            q_objects |= Q(bank_account__public_id__in=bank_ids)

        transactions = BimaTreasuryTransaction.objects.filter(q_objects)
        all_transactions = BimaTreasuryTransaction.objects.filter(q_objects).filter(date__lte=end_date)

        # if cash_ids:
        #     transactions = transactions.filter(cash__public_id__in=cash_ids)
        #     all_transactions = transactions.filter(cash__public_id__in=cash_ids)
        # if bank_ids:
        #     transactions = transactions.filter(bank_account__public_id__in=bank_ids)
        #     all_transactions = transactions.filter(bank_account__public_id__in=bank_ids)

        starting_balance = sum(
            t.amount if t.direction == TransactionDirection.INCOME.name else -t.amount
            for t in all_transactions.filter(date__lt=start_date)
        )

        month_year_keys = generate_month_year_keys(start_date, end_date)

        details = {key: {"total_income": 0, "total_outcome": 0, "difference": 0,
                         "starting_balance": 0, "closing_balance": 0, "increase_percentage": 0}
                   for key in month_year_keys}

        closing_balance = starting_balance

        total_income = 0
        total_outcome = 0

        for transaction in transactions.filter(date__range=(start_date, end_date)):
            key = f"{transaction.date.month}/{transaction.date.year}"

            if transaction.direction == TransactionDirection.INCOME.name:
                details[key]["total_income"] += transaction.amount
                total_income += transaction.amount
            else:
                details[key]["total_outcome"] += transaction.amount
                total_outcome += transaction.amount

        last_month_diff = None

        for key in month_year_keys:
            details[key]["difference"] = Bima360Helper.truncate(
                Decimal(details[key]["total_income"]) - Decimal(details[key]["total_outcome"]))
            details[key]["starting_balance"] = Bima360Helper.truncate(Decimal(closing_balance))
            closing_balance += details[key]["difference"]
            details[key]["closing_balance"] = Bima360Helper.truncate(Decimal(closing_balance))

            if last_month_diff is not None:
                if last_month_diff == 0:
                    details[key]["increase_percentage"] = 100 if details[key]["difference"] != 0 else 0
                else:
                    value = (Decimal(details[key]["difference"]) - Decimal(last_month_diff)) / Decimal(
                        last_month_diff) * 100
                    details[key]["increase_percentage"] = Bima360Helper.truncate(value)

            last_month_diff = Bima360Helper.truncate(Decimal(details[key]["difference"]))

        ordered_details = dict(
            sorted(details.items(), key=lambda x: (int(x[0].split('/')[1]), int(x[0].split('/')[0])))
        )

        difference = Bima360Helper.truncate(total_income - total_outcome)

        return {
            "starting_balance": starting_balance,
            "end_balance": closing_balance,
            "total_income": total_income,
            "total_outcome": total_outcome,
            "difference": difference,
            "details": ordered_details
        }


def generate_month_year_keys(start_date, end_date):
    keys = []
    curr_date = start_date
    while curr_date <= end_date:
        keys.append(f"{curr_date.month}/{curr_date.year}")
        curr_date += timedelta(days=(1 - curr_date.day) + 31)
        curr_date = curr_date.replace(day=1)
    return keys


class TransactionEffectStrategy:
    def apply(self, transaction):
        raise NotImplementedError

    def revert(self, transaction):
        raise NotImplementedError


class CashTransactionEffectStrategy(TransactionEffectStrategy):
    def apply(self, transaction):
        if transaction.direction == TransactionDirection.INCOME.name:
            transaction.cash.balance += transaction.amount
        else:
            transaction.cash.balance -= transaction.amount

        transaction.cash.save()

    def revert(self, transaction):
        if transaction.direction == TransactionDirection.INCOME.name:
            transaction.cash.balance -= transaction.amount
        else:
            transaction.cash.balance += transaction.amount
        transaction.cash.save()


class BankTransactionEffectStrategy(TransactionEffectStrategy):
    def apply(self, transaction):
        if transaction.direction == TransactionDirection.INCOME.name:
            transaction.bank_account.balance += transaction.amount
        else:
            transaction.bank_account.balance -= transaction.amount

        transaction.bank_account.save()

    def revert(self, transaction):
        if transaction.direction == TransactionDirection.INCOME.name:
            transaction.bank_account.balance -= transaction.amount
        else:
            transaction.bank_account.balance += transaction.amount

        transaction.bank_account.save()
