import csv
import io
import logging
from datetime import datetime
from uuid import UUID

import barcode
import numpy as np
import pandas as pd
import tablib
from PIL import Image
from PIL import ImageEnhance, ImageFilter
from barcode.writer import ImageWriter
from common.enums.global_enum import get_enum_value
from common.enums.product_enum import ENUM_MAPPINGS
from common.enums.product_enum import ProductType, PriceCalculationMethod, ProductStatus
from common.validators.file_validators import validate_file_extension_is_image, validate_file_size
from django.db import IntegrityError, transaction
from django.http import HttpResponse
from django.utils.translation import gettext_lazy as _
from erp.category.models import BimaErpCategory
from erp.unit_of_measure.models import BimaErpUnitOfMeasure
from erp.vat.models import BimaErpVat
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from pyzbar.pyzbar import decode
from rest_framework import status
from rest_framework.exceptions import ValidationError

from .models import BimaErpProduct

logger = logging.getLogger(__name__)


def generate_xls_file(data_to_export, model_fields):
    headers = [str(fd.verbose_name) for fd in model_fields.fields]
    file_data = tablib.Dataset()

    for product in data_to_export:
        values = []
        for fd in model_fields.fields:
            value = None
            try:
                if hasattr(getattr(product, fd.name), 'name'):
                    value = getattr(getattr(product, fd.name), 'name', None)
                elif fd.name in ENUM_MAPPINGS:
                    enum_value = getattr(product, fd.name, None)
                    value = str(ENUM_MAPPINGS[fd.name][enum_value].value) if enum_value is not None else \
                        getattr(product, fd.name, None)
                else:
                    value = getattr(product, fd.name, None)

                if isinstance(value, datetime):
                    value = value.replace(tzinfo=None)
                elif isinstance(value, UUID):
                    value = str(value)
            except Exception as ex:
                print(f"An error occurred while getting value of {fd.name}. Error: {str(ex)}")
            values.append(value)
        file_data.append(values)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=data.xlsx'

    workbook = Workbook()
    sheet = workbook.active

    # append headers
    sheet.append(headers)

    header_font = Font(size=12, bold=True)
    for cell in sheet[1]:
        cell.font = header_font

    for row_data in file_data:
        try:
            sheet.append(row_data)
        except ValueError as ex:
            print(ex)

    for column in range(1, len(headers) + 1):
        column_letter = get_column_letter(column)
        sheet.column_dimensions[column_letter].width = 12

    # Apply borders
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                    bottom=Side(style='thin'))
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.border = border

    quantity_index = headers.index(BimaErpProduct._meta.get_field('quantity').verbose_name)
    min_stock_level_index = headers.index(BimaErpProduct._meta.get_field('minimum_stock_level').verbose_name)
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row,
                               min_col=quantity_index + 1,
                               max_col=quantity_index + 1):
        for cell in row:
            corresponding_min_stock_level_cell = sheet.cell(row=cell.row, column=min_stock_level_index + 1)
            if cell.value is not None and corresponding_min_stock_level_cell.value is not None and cell.value <= corresponding_min_stock_level_cell.value:
                cell.fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')  # Red color
            else:
                cell.fill = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')  # Green color

    for column in sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    workbook.save(response)

    return response


def export_to_csv(products, model_fields):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="products.csv"'
    field_names_to_show = [fd.name for fd in model_fields.fields]
    writer = csv.writer(response)
    writer.writerow(field_names_to_show)

    for product in products:
        row_data = []
        for field in field_names_to_show:
            value = getattr(product, field)
            if hasattr(value, 'name'):
                value = getattr(value, 'name', None)
            elif field in ENUM_MAPPINGS:
                enum_value = getattr(product, field, None)
                value = ENUM_MAPPINGS[field][enum_value].value if enum_value is not None else None
            row_data.append(value if value is not None else '')
        writer.writerow(row_data)

    return response


def enhance_image(image):
    enhancer = ImageEnhance.Contrast(image)
    enhanced_image = enhancer.enhance(1.5)
    enhanced_image = enhanced_image.filter(ImageFilter.MedianFilter)
    return enhanced_image


def generate_file(barcode_file):
    response = HttpResponse(content_type="image/jpeg")
    barcode_file.save(response, "JPEG")
    return response


def read_barcode_from_image(image_array):
    decoded_objects = decode(image_array)
    if decoded_objects:
        return decoded_objects[0].data.decode('utf-8')
    else:
        return None


def verify_file_exist_for_ean13(request):
    barcode_image = request.FILES.get('barcode_image')
    if not barcode_image:
        raise ValidationError(_("Impossible de lire le fichier"))
    validate_file_extension_is_image(barcode_image)
    validate_file_size(barcode_image)
    return barcode_image


def generate_ean13_from_image(barcode_image):
    try:
        img_array = np.asarray(bytearray(barcode_image.read()), dtype=np.uint8)
        image = Image.open(io.BytesIO(img_array))

        if image.mode != 'RGB':
            image = image.convert('RGB')

        enhanced_image = enhance_image(image)
        enhanced_image = enhanced_image.convert('L')
        enhanced_np_array = np.array(enhanced_image)
        decoded_data = read_barcode_from_image(np.array(enhanced_np_array))

        if decoded_data:
            bar = barcode.get_barcode(name='code128', code=decoded_data, writer=ImageWriter())
            barcode_file = bar.render()
            generated_text = decoded_data

            return {'ean13': generated_text}, status.HTTP_200_OK

        else:
            raise Exception(_('Impossible de générer le code à barre'))

    except ValidationError as ex:
        logger.error(f'Validation error while generating barcode from file: {str(ex)}')
        return {'error': str(ex)}, status.HTTP_400_BAD_REQUEST

    except Exception as ex:
        logger.error(f'Unexpected error while generating barcode from file: {str(ex)}')
        return {'error': _('Impossible de générer le code à barre')}, status.HTTP_400_BAD_REQUEST


def verify_all_required_filed_exist(**kwargs):
    error_message = []
    required_fields = ['name', 'reference', 'type', 'price_calculation_method', 'status']
    for field in required_fields:
        if not kwargs.get(field):
            error_message.append({
                'error': _('{} does not exist').format(field.replace('_', ' ').capitalize()),
                'data': kwargs.get('name', "")
            })

    return len(error_message) > 0, error_message


def import_product_data_from_csv_file(df):
    error_rows = []
    created_count = 0

    for index, row in df.iterrows():
        try:
            name = str(row.get('name')) if pd.notnull(row.get('name')) else ""
            reference = str(row.get('reference')) if pd.notnull(row.get('reference')) else ""
            description = str(row.get('description')) if pd.notnull(row.get('description')) else ""
            ean13 = str(row.get('ean13')) if pd.notnull(row.get('ean13')) else ""
            type = get_enum_value(ProductType, str(row.get('type'))) if pd.notnull(row.get('type')) else None
            purchase_price = validate_decimal(row.get('purchase_price'), 'purchase_price')
            sell_price = validate_decimal(row.get('sell_price'), 'sell_price')
            price_calculation_method = get_enum_value(PriceCalculationMethod,
                                                      str(row.get('price_calculation_method'))) if pd.notnull(
                row.get('price_calculation_method')) else None
            sell_percentage = validate_decimal(row.get('sell_percentage'), 'sell_percentage')
            category_name = str(row.get('category')) if pd.notnull(row.get('category')) else ""
            vat_name = str(row.get('vat')) if pd.notnull(row.get('vat')) else ""
            unit_of_measure_name = str(row.get('unit_of_measure')) if pd.notnull(row.get('unit_of_measure')) else ""
            status = get_enum_value(ProductStatus, str(row.get('status'))) if pd.notnull(row.get('status')) else None
            minimum_stock_level = validate_integer(row.get('minimum_stock_level'), 'minimum_stock_level')
            maximum_stock_level = validate_integer(row.get('maximum_stock_level'), 'maximum_stock_level')
            dimension = str(row.get('dimension')) if pd.notnull(row.get('dimension')) else ""
            weight = str(row.get('weight')) if pd.notnull(row.get('weight')) else ""
            reorder_point = validate_integer(row.get('reorder_point'), 'reorder_point')
            lead_time = validate_integer(row.get('lead_time'), 'lead_time')
            serial_number = str(row.get('serial_number')) if pd.notnull(row.get('serial_number')) else ""
            quantity = validate_decimal(row.get('quantity'), 'quantity')
            virtual_quantity = validate_decimal(row.get('virtual_quantity'), 'virtual_quantity')

            exist_errors, error_message = verify_all_required_filed_exist(name, reference, type,
                                                                          price_calculation_method,
                                                                          status)
            if exist_errors:
                error_rows = error_message

            category = BimaErpCategory.objects.filter(name=category_name).first() if category_name else None
            vat = BimaErpVat.objects.filter(name=vat_name).first() if vat_name else None
            unit_of_measure = BimaErpUnitOfMeasure.objects.filter(
                name=unit_of_measure_name).first() if unit_of_measure_name else None

            if not category or not vat or not unit_of_measure:
                error_rows.append(
                    {'error': _('Category, VAT or Unit of Measure does not exist'), 'data': name if name else ""})
                continue

            with transaction.atomic():
                product = BimaErpProduct(
                    name=name,
                    reference=reference,
                    description=description,
                    ean13=ean13,
                    type=type,
                    purchase_price=purchase_price,
                    sell_price=sell_price,
                    price_calculation_method=price_calculation_method,
                    sell_percentage=sell_percentage,
                    category=category,
                    vat=vat,
                    unit_of_measure=unit_of_measure,
                    status=status,
                    minimum_stock_level=minimum_stock_level,
                    maximum_stock_level=maximum_stock_level,
                    dimension=dimension,
                    weight=weight,
                    reorder_point=reorder_point,
                    lead_time=lead_time,
                    serial_number=serial_number,
                    quantity=quantity,
                    virtual_quantity=virtual_quantity
                )
                product.full_clean()  # Validate the model instance
                product.save()
            created_count += 1
        except ValidationError as e:
            error_rows.append({'error': _('Invalid data: {}').format(e), 'data': name if name else ""})
        except IntegrityError as e:
            error_message = _('Integrity error occurred: {}').format(str(e))
            error_rows.append({'error': str(error_message), 'data': name if name else ""})
        except Exception as e:
            error_rows.append({'error': str(e), 'data': name if name else ""})

    return error_rows, created_count


def validate_decimal(value, field_name, default=0.0):
    try:
        return float(value) if pd.notnull(value) else default
    except ValueError:
        return default


def validate_integer(value, field_name, default=0):
    try:
        return int(value) if pd.notnull(value) else default
    except ValueError:
        return default
